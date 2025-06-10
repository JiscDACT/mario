import logging
from copy import copy
from openpyxl import load_workbook
from datetime import date
import pandas as pd
from pandas import DataFrame

from mario.data_extractor import DataExtractor
from mario.dataset_specification import DatasetSpecification
from mario.metadata import Metadata
from mario.options import ExcelOptions

logger = logging.getLogger(__name__)
style_attrs = ["alignment", "border", "fill", "font", "number_format", "protection"]


def rotate_list(list_1, n):
    """
    Rotates a list n positions. Used to push measure to the front of the data
    :param list_1: the list to rotate
    :param n: number of positions to rotate
    :return: rotated list
    """
    return list_1[-n:] + list_1[:-n]


class ExcelBuilder(object):
    """
    Class for creating Excel outputs
    """

    def __init__(self,
                 output_file_path: str,
                 data_extractor: DataExtractor,
                 dataset_specification: DatasetSpecification,
                 metadata: Metadata,
                 **kwargs
                 ):

        self.data_extractor = data_extractor
        self.dataset_specification = dataset_specification
        self.metadata = metadata
        self.filepath = output_file_path
        self.options = ExcelOptions(**kwargs)
        self.workbook = None
        self.rows = None
        self.cols = None
        self.total = 0
        logger.debug("Excel handler initialised")

    def create_workbook(self, create_notes_page=False):
        """
        Creates a write-only workbook and builds
        content in streaming mode to conserve memory
        """
        template_workbook = load_workbook(self.options.template_path)
        template_workbook.save(self.filepath)

        if create_notes_page:
            self.__create_notes_page__()
        else:
            self.__delete_notes_page__()
        self.__create_data_page__()
        self.__create_pivot_page__()

        # Set active sheet
        workbook = load_workbook(self.filepath)
        for sheet in workbook:
            workbook[sheet.title].views.sheetView[0].tabSelected = False
        workbook.save(self.filepath)
        workbook.close()

        logger.debug("Completed workbook")

    def __format_cell__(self, cell, new_cell):
        """
        Copy attributes from src to dst. Attributes are shallow-copied to avoid
        TypeError: unhashable type: 'StyleProxy'
        """
        for name in style_attrs:
            setattr(new_cell, name, copy(getattr(cell, name)))
        return new_cell

    def __replace_template_values(self, cell):
        """
        Replaces template values in a cell with metadata from the order
        """
        value = cell.value
        if value == '{ENQUIRY_NUMBER}':
            value = self.dataset_specification.collection
        if value == '{CUSTOMER_REFERENCE}':
            value = self.dataset_specification.get_property('customerRef')
        if value == '{USER}':
            value = 'Jisc Tailored Datasets App v1.0'
        if value == '{DATE}':
            value = date.today().isoformat()
        if value == '{DATASOURCE}':
            value = self.dataset_specification.get_property('datasource')
        if value == '{POP}':
            value = self.total
        if value == '{CAT}':
            value = self.dataset_specification.get_property('onwardUseCategory')
        if value == '{FORMAT}':
            value = 'Excel pivot table'
        return value

    def __create_notes_page__(self):
        workbook = load_workbook(self.filepath)
        notes = workbook.get_sheet_by_name('Notes')
        self.__update_notes__(notes)
        workbook.save(self.filepath)

    def __delete_notes_page__(self):
        workbook = load_workbook(self.filepath)
        if 'Notes' in workbook.sheetnames:
            workbook.remove(workbook['Notes'])
            workbook.save(self.filepath)

    def __create_pivot_page__(self):
        workbook = load_workbook(self.filepath)
        pivot = workbook['Pivot']._pivots[0]
        # Update the pivot table range
        pivot.cache.cacheSource.worksheetSource.ref = self.__range__()
        # Set to refresh
        pivot.cache.refreshOnLoad = True
        workbook.save(self.filepath)

    def create_notes_only(self, filename=None, data_format='CSV'):
        """
        Create a Notes page only, to accompany CSV outputs
        """
        self.workbook = load_workbook(self.options.template_path)
        self.__update_notes__(data_format=data_format, ws=self.workbook.get_sheet_by_name('Notes'))
        self.workbook.remove(self.workbook.get_sheet_by_name('Data'))
        self.workbook.remove(self.workbook.get_sheet_by_name('Pivot'))
        if filename is not None:
            self.workbook.save(filename=filename)
        else:
            self.workbook.save(filename=self.filepath)

    def __create_data_page__(self):

        if self.options.validate:
            if not self.data_extractor.validate_data(allow_nulls=self.options.allow_nulls):
                raise ValueError("Validation error")

        df: DataFrame = self.data_extractor.get_data_frame(
            minimise=self.options.minimise,
            include_row_numbers=self.options.include_row_numbers
        )

        # Reorder the columns to put the measure in col #1
        cols = df.columns.tolist()
        cols = rotate_list(cols, 1)
        df = df[cols]

        self.rows = len(df)
        self.cols = len(cols)

        with pd.ExcelWriter(
                self.filepath,
                engine='openpyxl',
                mode="a",
                if_sheet_exists='replace'
        ) as writer:
            df.to_excel(excel_writer=writer, sheet_name='Data', index=False)

    def __update_notes__(self, ws, data_format='Excel pivot table'):
        """
        Updates the notes page
        """
        self.total = self.data_extractor.get_total()

        ws['B6'] = self.dataset_specification.collection
        ws['B7'] = self.dataset_specification.get_property('customerRef')
        ws['B9'] = 'Jisc Tailored Datasets App v1.0'
        ws['B11'] = date.today().isoformat()
        ws['B14'] = self.dataset_specification.get_property('datasource')
        ws['B15'] = self.total
        ws['B16'] = self.dataset_specification.get_property('onwardUseCategory')
        ws['B17'] = data_format

        # Field names and definitions
        row = 32
        for item in self.dataset_specification.items:
            meta = self.metadata.get_metadata(item)
            ws['A' + str(row)] = item
            if meta.get_property('tdsaDescription'):
                ws['B' + str(row)] = meta.get_property('tdsaDescription')
            row += 1

    def __range__(self):
        """
        Determine the cell range based on the dataset size
        :return: a cell range e.g "A1:C26"
        """
        return 'A1' + ':' + chr(ord('@') + self.cols) + str(self.rows + 1)
