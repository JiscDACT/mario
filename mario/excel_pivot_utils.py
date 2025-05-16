from typing import List
from openpyxl import Workbook
import pandas as pd


def replace_pivot_cache_with_subset(worksheet, field, value) -> None:
    """
    Replaces the cached data for a pivot with a subset filtered by
    the supplied field and value
    :param worksheet: the sheet to filter
    :param field: the field name
    :param value: the field value to filter by
    :return: None
    """
    pivot = worksheet._pivots[0]
    cached_data = pivot.cache
    subset = get_subset(cached_data, field, value)
    pivot_cache = worksheet._pivots[0].cache
    pivot_cache.records.r = subset


def get_subset(pivot_cache, field_name, field_value):
    """
    Returns a subset from pivot cache data using the field
    name and field value as a filter
    :param pivot_cache: the cache to filter
    :param field_name: the name of the field to filter by
    :param field_value: the value to filter by
    :return: a filtered pivot cache
    """
    # Get the index of the field name
    field_index = None
    cache_field = None
    field_value_index = None

    for idx, field in enumerate(pivot_cache.cacheFields):
        if field.name == field_name:
            cache_field = field
            field_index = idx
            break

    # Get the index of the field value
    # Map the indices to actual values
    field_items = cache_field.sharedItems._fields
    for index, item in enumerate(field_items):
        if item.v == field_value:
            field_value_index = index
            break

    if field_index is None:
        raise ValueError(f"Field '{field_name}' not found in pivot cache.")
    if field_value_index is None:
        raise ValueError(f"Value '{field_value}' not found in pivot cache.")

    # Filter the pivot cache records
    filtered_records = []
    for record in pivot_cache.records.r:
        if record._fields[field_index].v == field_value_index:
            filtered_records.append(record)

    return filtered_records


def get_unique_values_from_pivot_worksheet(workbook, sheet_name, field) -> List[str]:
    ws = workbook[sheet_name]
    pivot = ws._pivots[0]
    cached_data = pivot.cache

    items = []

    for f in cached_data.cacheFields:
        if f.name == field:
            for item in f.sharedItems._fields:
                if hasattr(item, 'v'):
                    items.append(item.v)
                elif hasattr(item, 'n'):
                    items.append(item.n)
                elif hasattr(item, 'f'):
                    items.append(item.f)
                else:
                    items.append(item)

    if len(items) == 0:
        raise ValueError(f"Column '{field}' not found in pivot_data")

    # Use a set to get unique values
    unique_values = set(items)
    return list(unique_values)


def get_unique_values_from_pivot_cache(file_path, sheet_name, field) -> List[str]:
    """
    Returns all the unique values for a column in a pivot cache
    :param file_path: the workbook file path
    :param sheet_name: the name of the sheet
    :param field: the field to get values for
    :return: a list of unique values for the specified column in the cache
    """
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True)
    return get_unique_values_from_pivot_worksheet(workbook=wb, sheet_name=sheet_name, field=field)


def get_header(file_path: str, sheet_name: str):
    header: List[str] = pd.read_excel(file_path, sheet_name=sheet_name, nrows=0).columns.tolist()
    return header


def sheet_contains_column_name(file_path: str, sheet_name: str, field: str):
    return field in get_header(file_path=file_path, sheet_name=sheet_name)


def get_unique_values_for_workbook(file_path, field) -> List[str] | None:
    """
    Gets the unique values for a field from an Excel workbook. The
    function iterates over the sheets in the workbook until it
    finds a sheet containing the data - either a regular sheet or
    a sheet containing a pivot.
    :param file_path: the path to the Excel file
    :param field: the field to get values for
    :return: a list of unique values
    :raises: ValueError if no sheets contain the field
    """
    from openpyxl import load_workbook
    wb: Workbook = load_workbook(file_path, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        if len(ws._pivots) > 0:
            return get_unique_values_from_pivot_worksheet(workbook=wb, sheet_name=sheet, field=field)
        else:
            if sheet_contains_column_name(file_path=file_path, sheet_name=sheet, field=field):
                return pd.read_excel(file_path, sheet_name=sheet, dtype={field: object}, usecols=[field])[field].unique().tolist()
    raise ValueError("No valid worksheet containing field values")


def get_info_sheets(workbook, field):
    info_sheets = []
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        if len(ws._pivots) == 0:
            headers = [c.value for c in next(workbook[sheet].iter_rows(min_row=1, max_row=1))]
            if field not in headers:
                info_sheets.append(sheet)
    return info_sheets


def set_excel_active_sheet(file_path: str):
    """
    Deselects the active tabs, forcing the workbook to refresh on open.
    Useful for refreshing pivots and ensuring tabs aren't grouped on open
    :param file_path: the path to the workbook
    :return: None
    """
    from openpyxl import load_workbook

    workbook: Workbook = load_workbook(file_path)
    for sheet in workbook:
        workbook[sheet.title].views.sheetView[0].tabSelected = False
    workbook.save(file_path)
    workbook.close()


def get_worksheet_containing_field(workbook, field):
    for sheet in workbook.sheetnames:
        headers = [c.value for c in next(workbook[sheet].iter_rows(min_row=1, max_row=1))]
        if field in headers:
            return sheet


def prepend_sheet_to_workbook(source_workbook_file, target_workbook_file, sheet_name) -> None:
    """
    Prepends the specified sheet from the source workbook to the target workbook
    :param source_workbook_file: the source Excel file
    :param target_workbook_file: the target Excep file
    :param sheet_name: the sheet to prepend
    :return: None
    """
    from openpyxl import Workbook, load_workbook
    target_workbook: Workbook = load_workbook(target_workbook_file)
    source_workbook = load_workbook(source_workbook_file)
    source = source_workbook[sheet_name]
    source._parent = target_workbook
    target_workbook._add_sheet(sheet=source)
    target_workbook.move_sheet(sheet=source, offset=-(len(target_workbook.sheetnames) - 1))
    target_workbook.save(target_workbook_file)



