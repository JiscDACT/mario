"""
A utility for partitioning dataset files by a particular field
"""
import os
import shutil
import csv
import logging

logger = logging.getLogger(__name__)


class DatasetSplitter:

    def __init__(self,
                 field: str,
                 source_path: str,
                 output_path: str,
                 use_excel_builder=False,
                 template=None,
                 append_info_sheets=False
                 ):
        self.field = field
        self.source_path = source_path
        self.output_path = output_path
        self.other_files = []
        self.use_excel_builder = use_excel_builder
        self.template = template
        self.append_info_sheets = append_info_sheets

        if self.use_excel_builder is True and (template is None or not os.path.exists(template)):
            raise ValueError("You must specify a valid template file to use the Excel Builder")

        # If the source path doesn't exist, raise an error
        if not os.path.exists(self.source_path):
            raise FileNotFoundError("Source folder doesn't exist.")

        # If the source path and output path are the same, raise an error
        if self.source_path == self.output_path:
            raise ValueError("Source and output folders can't be the same.")

        # If the output path contains subdirectories, raise an error
        if os.path.exists(self.output_path):
            if any(os.path.relpath(root, self.output_path).count(os.sep) > 0 for root, dirs, _ in os.walk(self.output_path)):
                raise ValueError("Output folder contains subdirectories so is probably not something you "
                                 "want to delete and overwrite.")

        # If the source path contains subdirectories, raise an error
        if next(os.walk(self.source_path))[1]:
            raise ValueError("Source folder contains subdirectories so is probably not something you "
                             "want to split as the subdirectories aren't going to be processed.")

        # Delete the output path if it exists
        if os.path.exists(self.output_path):
            logger.info(f"Removing output path at {self.output_path}")
            shutil.rmtree(self.output_path)

        os.makedirs(self.output_path)

    def get_output_path(self, file, field_value):
        return os.path.join(self.output_path, field_value, file)

    def get_files_to_split_or_copy(self):
        # Iterate over each file in the source folder
        files_to_split = []
        files_to_copy = []
        for file in os.listdir(self.source_path):
            if os.path.isfile(os.path.join(self.source_path, file)):
                if file.endswith('.xlsx'):
                    try:
                        self.get_excel_values(file)
                        files_to_split.append(file)
                        logger.debug(f"Excel file {file} is to be split")
                    except Exception:
                        logger.debug(f"Excel file {file} has no values for field")
                        files_to_copy.append(file)
                elif file.endswith('.csv'):
                    files_to_split.append(file)
                elif file.endswith('.csv.gz'):
                    files_to_split.append(file)
                else:
                    files_to_copy.append(file)
        if len(files_to_split) == 0:
            raise ValueError("No files were split")
        return files_to_split, files_to_copy

    def split_file(self, file):
        if file.endswith('.xlsx'):
            if self.use_excel_builder:
                self.split_excel_using_builder(file)
            else:
                self.split_excel(file)
        elif file.endswith('.csv'):
            self.split_csv(file)
        elif file.endswith('.csv.gz'):
            self.split_gzipped_csv(file)

    def split_files(self):
        files_to_split, files_to_copy = self.get_files_to_split_or_copy()
        for file in files_to_split:
            self.split_file(file)

    def copy_other_file(self, file):
        logger.info(f"Copying file {file} to output directory")
        for output_directory in os.listdir(self.output_path):
            if os.path.isdir(os.path.join(self.output_path, output_directory)):
                shutil.copyfile(
                    os.path.join(self.source_path, file),
                    os.path.join(self.output_path, output_directory, file))

    def copy_other_files(self) -> None:
        """
        Iterates over the source folder, and copies each non-data file into
        each 'split' output directory present
        :return: None
        """
        files_to_split, files_to_copy = self.get_files_to_split_or_copy()
        for file in files_to_copy:
            self.copy_other_file(file)

    def process_batch(self, batch, column_name, file_handles, file_name):
        for row in batch:
            value = row[column_name]
            if value not in file_handles:
                os.makedirs(os.path.join(self.output_path, value), exist_ok=True)
                file_path = self.get_output_path(field_value=value, file=file_name)
                file_handles[value] = open(file_path, 'w', newline='')
                writer = csv.DictWriter(file_handles[value], fieldnames=row.keys())
                writer.writeheader()
            writer = csv.DictWriter(file_handles[value], fieldnames=row.keys())
            writer.writerow(row)

    def split_gzipped_csv(self, file_name: str, batch_size=10000):
        import gzip

        logger.info(f"Splitting gzipped CSV file {file_name}")

        # Dict for holding open file handles
        file_handles = {}

        # Path to the CSV to split
        file_path = os.path.join(self.source_path, file_name)

        output_file_name = file_name.rstrip('.gz')

        with gzip.open(file_path, 'rt', newline='') as infile:
            reader = csv.DictReader(infile)

            # Process the input file in batches
            batch = []
            for row in reader:
                batch.append(row)
                if len(batch) >= batch_size:
                    self.process_batch(batch, self.field, file_handles, output_file_name)
                    batch = []

            # Process any remaining rows in the last batch
            if batch:
                self.process_batch(batch, self.field, file_handles, output_file_name)

        # Close all open file handles
        for handle in file_handles.values():
            handle.close()

    def split_csv(self, file_name: str, batch_size=10000, compression=None):

        logger.info(f"Splitting CSV file {file_name}")

        # Dict for holding open file handles
        file_handles = {}

        # Path to the CSV to split
        file_path = os.path.join(self.source_path, file_name)

        # Open the input CSV file for reading
        with open(file_path, 'r') as infile:
            reader = csv.DictReader(infile)

            # Process the input file in batches
            batch = []
            for row in reader:
                batch.append(row)
                if len(batch) >= batch_size:
                    self.process_batch(batch, self.field, file_handles, file_name)
                    batch = []

            # Process any remaining rows in the last batch
            if batch:
                self.process_batch(batch, self.field, file_handles, file_name)

        # Close all open file handles
        for handle in file_handles.values():
            handle.close()

    def split_excel_using_builder(self, file_name: str) -> None:
        """
        Splits an Excel file by using the ExcelBuilder to create
        a new workbook containing both a Data sheet and a Pivot sheet
        :param file_name: the name of the Excel file to split
        :return: None
        """
        from mario.excel_builder import ExcelBuilder
        from mario.data_extractor import DataFrameExtractor
        from mario.dataset_specification import DatasetSpecification
        from mario.metadata import Metadata, Item
        from mario.excel_pivot_utils import prepend_sheet_to_workbook, get_worksheet_containing_field, get_info_sheets
        import pandas as pd
        from openpyxl import load_workbook

        file_path = os.path.join(self.source_path, file_name)

        # Create a dataset based on the contents of the 'Data' sheet in the Excel workbook. We find this
        # by locating a 'data sheet' that has the field name in the header. At the same time we identify
        # any 'info' sheets that don't have any data as we want to copy those as-is
        workbook = load_workbook(file_path)
        data_sheet = get_worksheet_containing_field(workbook, self.field)

        if data_sheet is None:
            logger.warning("Encountered an Excel file with no data sheet; treating as an 'other' file")
            self.other_files.append(file_name)
            return

        # Read in the data and get unique values for the field we want to split by
        df = pd.read_excel(file_path, sheet_name=data_sheet,  dtype={self.field: object}, header=0)
        values = df[self.field].unique()

        # Generate some basic metadata
        ds = DatasetSpecification()
        ds.dimensions = list(df.columns)
        ds.measures = []
        md = Metadata()
        for field in df.columns:
            item = Item()
            item.name = field
            md.add_item(item)

        # Create a data extractor instance for the DF
        dx = DataFrameExtractor(
            dataset_specification=ds,
            metadata=md,
            dataframe=df
        )

        # Create an ExcelBuilder instance for creating excel files
        builder = ExcelBuilder(
            output_file_path='',
            template_path=self.template,
            dataset_specification=ds,
            metadata=md,
            data_extractor=dx
        )

        # For each unique value, subset the data and create the output
        for value in values:
            logger.info(f"Splitting {file_name} for value {value}")
            output_folder = os.path.join(self.output_path, value)
            os.makedirs(output_folder, exist_ok=True)
            output_path = os.path.join(output_folder, file_name)
            builder.filepath = output_path
            subset = df[df[self.field] == value]
            dx._data = subset
            builder.create_workbook(create_notes_page=False)

            # prepend any info sheets, if present
            if self.append_info_sheets:
                for info_sheet in get_info_sheets(workbook, self.field):
                    prepend_sheet_to_workbook(
                        source_workbook_file=file_path,
                        target_workbook_file=output_path,
                        sheet_name=info_sheet
                    )

    def get_excel_values(self, file_name: str):
        from mario.excel_pivot_utils import get_unique_values_for_workbook
        file_path = os.path.join(self.source_path, file_name)
        return get_unique_values_for_workbook(file_path=file_path, field=self.field)

    def split_excel_by_value(self, file_name, value: str):
        """
        Splits an Excel file by a single value.
        :param file_name: the name of the Excel file (i.e. relative to the source path)
        :param value: the value to split by
        :return: the path to the split file for the value given e.g. output/Central/filename.xlsx
        """
        import pandas as pd
        from openpyxl import load_workbook
        from mario.excel_pivot_utils import set_excel_active_sheet

        file_path = os.path.join(self.source_path, file_name)

        logger.info(f"Splitting for value {value}")
        sheet_names = pd.ExcelFile(file_path).sheet_names
        split_output_path = os.path.join(self.output_path, str(value))
        split_workbook_path = os.path.join(split_output_path, file_name)
        os.makedirs(split_output_path, exist_ok=True)
        shutil.copyfile(src=file_path, dst=split_workbook_path)

        # Load the workbook once per 'run'
        workbook = load_workbook(split_workbook_path)
        for sheet_name in sheet_names:
            if len(workbook.get_sheet_by_name(sheet_name)._pivots) > 0:
                logger.info(f"Splitting excel pivot")
                self.split_excel_pivot(
                    workbook=workbook,
                    file_path=split_workbook_path,
                    sheet_name=sheet_name,
                    value=value
                )
            else:
                logger.info(f"Splitting excel sheet")
                self.split_excel_table(
                    workbook=workbook,
                    file_path=split_workbook_path,
                    sheet_name=sheet_name,
                    value=value
                )
        # Close the workbook explicitly once we're finished with it

        workbook.close()
        # Clears the active sheet so we don't have any grouped tabs. This
        # also seems to have a side effect of refreshing the pivot cache
        set_excel_active_sheet(split_workbook_path)

        return split_workbook_path

    def split_excel(self, file_name: str):
        """
        Convenience method for splitting an Excel file. Call the split methods
        directly to process in parallel.
        :param file_name:
        :return:
        """
        from mario.excel_pivot_utils import set_excel_active_sheet
        # Get the values
        try:
            values = self.get_excel_values(file_name=file_name)
        except ValueError:
            logger.warning("Encountered an Excel file with no data; treating as an 'other' file")
            self.other_files.append(file_name)
            return False

        # Create split workbooks
        for value in values:
            split_workbook_path = self.split_excel_by_value(file_name=file_name, value=value)
            set_excel_active_sheet(file_path=split_workbook_path)

    def split_excel_pivot(self, workbook, file_path: str, sheet_name: str, value: str):
        """
        Splits an Excel pivot sheet
        :param workbook: the workbook to split
        :param file_path: the path to write the split file to
        :param sheet_name: the name of the sheet
        :param value: the value to split by
        :return: None
        """
        from mario.excel_pivot_utils import replace_pivot_cache_with_subset
        ws = workbook[sheet_name]
        replace_pivot_cache_with_subset(ws, self.field, value)
        workbook.save(file_path)

    def split_excel_table(self, workbook, file_path: str, sheet_name: str, value):
        """
        Splits a standard worksheet (i.e. not a pivot) for the specified value
        :param workbook: the workbook to split
        :param file_path: the file path to write the split file to
        :param sheet_name: the name of the sheet
        :param value: the value to split by
        :return: None
        """

        ws = workbook[sheet_name]
        header = [cell.value for cell in ws[1]]  # Read header row

        if self.field not in header:
            return None

        # Create a new sheet for filtered data
        workbook.create_sheet(title=f"{sheet_name}_filtered")
        new_ws = workbook[f"{sheet_name}_filtered"]
        new_ws.append(header)

        # Read rows efficiently. Note that using comprehension first rather than writing direct from iter_rows
        # speeds things up by about 50% at a cost of higher memory use.
        field_index = header.index(self.field)
        filtered_rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if row[field_index] == value]
        for row in filtered_rows:
            new_ws.append(row)

        # Save and remove original sheet
        workbook.remove(ws)
        new_ws.title = sheet_name
        workbook.save(file_path)
