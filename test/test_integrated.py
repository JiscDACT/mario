import os

import pandas as pd
import pytest

from mario.data_extractor import Configuration, HyperFile, DataExtractor, StreamingDataExtractor
from mario.dataset_builder import DatasetBuilder, Format
from mario.dataset_specification import dataset_from_json, Constraint, dataset_from_manifest
from mario.metadata import metadata_from_json, metadata_from_manifest
from mario.query_builder import SubsetQueryBuilder, ViewBasedQueryBuilder


def test_integration_tdsx():
    dataset = dataset_from_json(os.path.join('test', 'dataset.json'))
    metadata = metadata_from_json(os.path.join('test', 'metadata.json'))
    configuration = Configuration(file_path=os.path.join('test', 'orders.hyper'))
    extractor = HyperFile(configuration=configuration, dataset_specification=dataset, metadata=metadata)
    builder = DatasetBuilder(dataset_specification=dataset, metadata=metadata, data=extractor)
    path = os.path.join('output', dataset.collection, dataset.name + '.tdsx')
    os.makedirs(os.path.join('output', dataset.collection), exist_ok=True)
    builder.build(file_path=path, output_format=Format.TABLEAU_PACKAGED_DATASOURCE)


def test_integration_csv():
    dataset = dataset_from_json(os.path.join('test', 'dataset.json'))
    metadata = metadata_from_json(os.path.join('test', 'metadata.json'))
    configuration = Configuration(file_path=os.path.join('test', 'orders.hyper'))
    extractor = HyperFile(configuration=configuration, dataset_specification=dataset, metadata=metadata)
    builder = DatasetBuilder(dataset_specification=dataset, metadata=metadata, data=extractor)
    path = os.path.join('output', dataset.collection, dataset.name + '.csv')
    os.makedirs(os.path.join('output', dataset.collection), exist_ok=True)
    builder.build(file_path=path, output_format=Format.CSV)


def test_integration_excel():
    dataset = dataset_from_manifest(os.path.join('test', 'manifest_superstore.json'))
    metadata = metadata_from_manifest(os.path.join('test', 'manifest_superstore.json'))
    configuration = Configuration(file_path=os.path.join('test', 'orders.hyper'))
    extractor = DataExtractor(configuration=configuration, dataset_specification=dataset, metadata=metadata)
    builder = DatasetBuilder(dataset_specification=dataset, metadata=metadata, data=extractor)
    path = os.path.join('output', 'test_integration_excel', 'data.xlsx')
    os.makedirs(os.path.join('output', 'test_integration_excel'), exist_ok=True)
    builder.build(file_path=path, output_format=Format.EXCEL_PIVOT, template_path='excel_template.xlsx')


def test_integration_excel_no_notes():
    from openpyxl import load_workbook
    dataset = dataset_from_manifest(os.path.join('test', 'manifest_superstore.json'))
    metadata = metadata_from_manifest(os.path.join('test', 'manifest_superstore.json'))
    configuration = Configuration(file_path=os.path.join('test', 'orders.hyper'))
    extractor = DataExtractor(configuration=configuration, dataset_specification=dataset, metadata=metadata)
    builder = DatasetBuilder(dataset_specification=dataset, metadata=metadata, data=extractor)
    path = os.path.join('output', 'test_integration_excel_no_notes', 'data.xlsx')
    os.makedirs(os.path.join('output', 'test_integration_excel_no_notes'), exist_ok=True)
    builder.build(file_path=path, output_format=Format.EXCEL_PIVOT, template_path='excel_template.xlsx')

    workbook = load_workbook(filename=path)
    assert 'Notes' not in workbook.sheetnames


def test_integration_excel_info_only():
    dataset = dataset_from_manifest(os.path.join('test', 'manifest_superstore.json'))
    metadata = metadata_from_manifest(os.path.join('test', 'manifest_superstore.json'))
    configuration = Configuration(file_path=os.path.join('test', 'orders.hyper'))
    extractor = DataExtractor(configuration=configuration, dataset_specification=dataset, metadata=metadata)
    builder = DatasetBuilder(dataset_specification=dataset, metadata=metadata, data=extractor)
    path = os.path.join('output', 'test_integration_excel_info', 'info.xlsx')
    os.makedirs(os.path.join('output', 'test_integration_excel_info'), exist_ok=True)
    builder.build(file_path=path, output_format=Format.EXCEL_INFO_SHEET, template_path='excel_template.xlsx')


def test_sql_extraction():
    # Skip this test if we don't have a connection string
    if not os.environ.get('CONNECTION_STRING'):
        pytest.skip("Skipping SQL test as no database configured")

    # Set up local test database, drivers and connection string to run this
    dataset = dataset_from_json(os.path.join('test', 'dataset.json'))
    metadata = metadata_from_json(os.path.join('test', 'metadata.json'))
    # Add constraints so we test parameter generation
    constraint = Constraint()
    constraint.item = 'Ship Mode'
    constraint.allowed_values = ['First Class']
    dataset.constraints.append(constraint)
    constraint = Constraint()
    constraint.item = 'City'
    constraint.allowed_values = ['Dallas']
    dataset.constraints.append(constraint)

    configuration = Configuration(
        connection_string=os.environ.get("CONNECTION_STRING"),
        schema='dev',
        view='superstore',
        query_builder=SubsetQueryBuilder
    )
    extractor = DataExtractor(
        dataset_specification=dataset,
        metadata=metadata,
        configuration=configuration
    )
    extractor.validate_data()
    sql_path = os.path.join('output', 'test_sql_extraction', 'query.sql')
    csv_path = os.path.join('output', 'test_sql_extraction', 'data.csv')
    os.makedirs(os.path.join('output', 'test_sql_extraction'), exist_ok=True)
    extractor.save_data_as_csv(csv_path)
    extractor.save_query(sql_path)


def test_sql_extraction_using_manifest():
    # Skip this test if we don't have a connection string
    if not os.environ.get('CONNECTION_STRING'):
        pytest.skip("Skipping SQL test as no database configured")

    # Set up local test database, drivers and connection string to run this
    dataset = dataset_from_manifest(os.path.join('test', 'manifest_superstore.json'))
    metadata = metadata_from_manifest(os.path.join('test', 'manifest_superstore.json'))
    configuration = Configuration(
        connection_string=os.environ.get("CONNECTION_STRING"),
        schema='dev',
        view='superstore',
        query_builder=SubsetQueryBuilder
    )
    extractor = DataExtractor(
        dataset_specification=dataset,
        metadata=metadata,
        configuration=configuration
    )
    extractor.validate_data()
    sql_path = os.path.join('output', 'test_sql_extraction_using_manifest', 'query.sql')
    csv_path = os.path.join('output', 'test_sql_extraction_using_manifest', 'data.csv')
    os.makedirs(os.path.join('output', 'test_sql_extraction_using_manifest'), exist_ok=True)
    extractor.save_data_as_csv(csv_path)
    extractor.save_query(sql_path)


def test_integration_excel_info_only_with_totals():
    from openpyxl import Workbook, load_workbook
    dataset = dataset_from_manifest(os.path.join('test', 'manifest_superstore.json'))
    metadata = metadata_from_manifest(os.path.join('test', 'manifest_superstore.json'))
    configuration = Configuration(file_path=os.path.join('test', 'orders.hyper'))
    extractor = DataExtractor(configuration=configuration, dataset_specification=dataset, metadata=metadata)
    total_from_query = extractor.get_total()
    builder = DatasetBuilder(dataset_specification=dataset, metadata=metadata, data=extractor)
    path = os.path.join('output', 'test_integration_excel_info', 'info.xlsx')
    os.makedirs(os.path.join('output', 'test_integration_excel_info'), exist_ok=True)
    builder.build(file_path=path, output_format=Format.EXCEL_INFO_SHEET, template_path='excel_template.xlsx')
    workbook: Workbook = load_workbook(path)
    total_from_notes = workbook.get_sheet_by_name('Notes')['B15'].value
    assert total_from_notes == total_from_query


def test_integration_csv_streaming():
    # Skip this test if we don't have a connection string
    if not os.environ.get('CONNECTION_STRING'):
        pytest.skip("Skipping SQL test as no database configured")
    dataset = dataset_from_json(os.path.join('test', 'dataset.json'))
    dataset.collection = 'test_integration_csv_streaming'
    metadata = metadata_from_json(os.path.join('test', 'metadata.json'))
    configuration = Configuration(
        connection_string=os.environ.get('CONNECTION_STRING'),
        schema="dev",
        view="superstore",
        query_builder=ViewBasedQueryBuilder
    )
    extractor = StreamingDataExtractor(
        dataset_specification=dataset,
        metadata=metadata,
        configuration=configuration
    )
    builder = DatasetBuilder(dataset_specification=dataset, metadata=metadata, data=extractor)
    path = os.path.join('output', dataset.collection, dataset.name + '.csv')
    os.makedirs(os.path.join('output', dataset.collection), exist_ok=True)
    builder.build(file_path=path, output_format=Format.CSV)
    df = pd.read_csv(path)
    assert len(df) == 10194


def test_integration_tdsx_streaming():
    # Skip this test if we don't have a connection string
    if not os.environ.get('CONNECTION_STRING'):
        pytest.skip("Skipping SQL test as no database configured")
    dataset = dataset_from_json(os.path.join('test', 'dataset.json'))
    dataset.collection = 'test_integration_tdsx_streaming'
    metadata = metadata_from_json(os.path.join('test', 'metadata.json'))
    configuration = Configuration(
        connection_string=os.environ.get('CONNECTION_STRING'),
        schema="dev",
        view="superstore",
        query_builder=ViewBasedQueryBuilder
    )
    extractor = StreamingDataExtractor(
        dataset_specification=dataset,
        metadata=metadata,
        configuration=configuration
    )
    builder = DatasetBuilder(dataset_specification=dataset, metadata=metadata, data=extractor)
    path = os.path.join('output', dataset.collection, dataset.name + '.tdsx')
    os.makedirs(os.path.join('output', dataset.collection), exist_ok=True)
    builder.build(file_path=path, output_format=Format.TABLEAU_PACKAGED_DATASOURCE)


def test_remove_redundant_hierarchies():
    metadata_file = os.path.join('test', 'metadata.json')
    metadata = metadata_from_json(file_path=metadata_file)
    dataset = dataset_from_json(os.path.join('test', 'dataset.json'))
    dataset.collection = 'test_remove_redundant_hierarchies'

    # remove an item so we have a one-item hierarchy for "Product"
    dataset.dimensions.remove("Product Name")
    assert "Product Name" not in dataset.items

    configuration = Configuration(file_path=os.path.join('test', 'orders.hyper'))
    extractor = HyperFile(configuration=configuration, dataset_specification=dataset, metadata=metadata)
    builder = DatasetBuilder(dataset_specification=dataset, metadata=metadata, data=extractor)

    # remove the product hierarchy
    builder.remove_redundant_hierarchies()

    assert "Product" not in metadata.get_hierarchies()
    assert "Location" in metadata.get_hierarchies()

    path = os.path.join('output', dataset.collection, dataset.name + '.tdsx')
    os.makedirs(os.path.join('output', dataset.collection), exist_ok=True)
    builder.build(file_path=path, output_format=Format.TABLEAU_PACKAGED_DATASOURCE)


def test_hyper_with_csv_output():
    folder = os.path.join('output', 'test_integration_hyper_with_csv_output')
    os.makedirs(folder, exist_ok=True)
    dataset = dataset_from_json(os.path.join('test', 'dataset.json'))
    metadata = metadata_from_json(os.path.join('test', 'metadata.json'))
    configuration = Configuration(
        file_path=os.path.join('test', 'orders.hyper')
    )
    extractor = HyperFile(
        dataset_specification=dataset,
        metadata=metadata,
        configuration=configuration
    )
    builder = DatasetBuilder(
        dataset_specification=dataset,
        metadata=metadata,
        data=extractor
    )
    builder.build(Format.CSV, file_path=os.path.join(folder, 'test.csv'))
    builder.build(Format.CSV, file_path=os.path.join(folder, 'test.csv.gz'))
    builder.build(Format.TABLEAU_PACKAGED_DATASOURCE, file_path=os.path.join(folder, 'test.tdsx'))
    builder.build(Format.EXCEL_PIVOT, file_path=os.path.join(folder, 'test.xlsx'))


def test_hyper_with_compressed_csv_output():
    folder = os.path.join('output', 'test_integration_hyper_with_csv_output')
    os.makedirs(folder, exist_ok=True)
    dataset = dataset_from_json(os.path.join('test', 'dataset.json'))
    metadata = metadata_from_json(os.path.join('test', 'metadata.json'))
    configuration = Configuration(
        file_path=os.path.join('test', 'orders.hyper')
    )
    extractor = HyperFile(
        dataset_specification=dataset,
        metadata=metadata,
        configuration=configuration
    )
    builder = DatasetBuilder(
        dataset_specification=dataset,
        metadata=metadata,
        data=extractor
    )
    builder.build(Format.CSV, file_path=os.path.join(folder, 'test.csv'))
    builder.data.save_data_as_csv(file_path=os.path.join(folder, 'test.csv'), compress_using_gzip=True)
    assert os.path.exists(os.path.join(folder, 'test.csv.gz'))
